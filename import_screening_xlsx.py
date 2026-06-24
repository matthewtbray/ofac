#!/usr/bin/env python3
"""
import_screening_xlsx.py  --  Import a screening list xlsx into ScreeningInput

Supports two file formats (auto-detected by header row):

  1. Delaware OFAC List (legacy format)
       Fixed layout, data starts at row 12, "As of" date in row 3.
       Blob auto-discovery: scans for "Delaware OFAC List-*.xlsx".

  2. Salesforce Report (SFDC format)
       Header row at row 1, data from row 2.
       Detected when cell A1 == "Contact: Contact ID".
       Must be supplied via --input (no blob auto-discovery for this format).

Column mapping — SFDC format:
  A  Contact: Contact ID        -> Contact_ID
  B  Subject Entity ID          -> Entity_ID
  C  Name                       -> ORG_NAME
  D  Contact: First Name        -> FIRST_NAME
  E  Contact: Last Name         -> LAST_NAME
  G  Contact: Mailing Street    -> ADDRESS1
  H  Contact: Mailing City      -> CITY
  I  Contact: Mailing State/Province -> STATE
  J  Contact: Mailing Zip/Postal Code -> POSTAL_CODE
  K  Contact: Mailing Country   -> COUNTRY

Usage
-----
  # SFDC local file
  python import_screening_xlsx.py --input "report1779464025538.xlsx" \\
      --server myserver.database.windows.net --database SDN [--drop]

  # Delaware OFAC List local file
  python import_screening_xlsx.py --input "Delaware OFAC List-2026-05-19.xlsx" \\
      --server myserver.database.windows.net --database SDN [--drop]

  # Auto-discover Delaware OFAC List from blob (Azure pipeline)
  python import_screening_xlsx.py \\
      --server myserver.database.windows.net --database SDN --drop

Environment variables:
  SQL_SERVER                Azure SQL logical server FQDN
  SQL_USER                  SQL login name
  SQL_PASSWORD              SQL password
  STORAGE_CONNECTION_STRING Required when --input is omitted
"""
import argparse
import os
import re
import sys
import tempfile
from datetime import datetime

import openpyxl
import pyodbc

# ---------------------------------------------------------------------------
# Delaware OFAC List layout constants  (1-indexed column numbers)
# ---------------------------------------------------------------------------
_DEL_DATA_START = 12
_DEL_AS_OF_ROW  = 3
_DEL_AS_OF_COL  = 2   # Column B
_DEL_ORG_COL    = 2   # Column B  -- entity / org name
_DEL_FNAME_COL  = 5   # Column E  -- Contact: First Name
_DEL_LNAME_COL  = 6   # Column F  -- Contact: Last Name
_DEL_ADDR_COL   = 8   # Column H  -- Contact: Mailing Address

# ---------------------------------------------------------------------------
# SFDC Report layout constants  (1-indexed column numbers)
# ---------------------------------------------------------------------------
_SFDC_HEADER_ROW   = 1
_SFDC_DATA_START   = 2
_SFDC_CONTACT_ID   = 1   # A
_SFDC_ENTITY_ID    = 2   # B
_SFDC_ORG_NAME     = 3   # C
_SFDC_FIRST_NAME   = 4   # D
_SFDC_LAST_NAME    = 5   # E
# Column F (6) = Contact: Mailing Address (full combined — not imported)
_SFDC_STREET       = 7   # G
_SFDC_CITY         = 8   # H
_SFDC_STATE        = 9   # I
_SFDC_POSTAL       = 10  # J
_SFDC_COUNTRY      = 11  # K

_SFDC_SENTINEL = 'Contact: Contact ID'   # value in A1 that identifies SFDC format

# ---------------------------------------------------------------------------
# Blob helpers
# ---------------------------------------------------------------------------
_BLOB_NAME_RE = re.compile(
    r'^Delaware OFAC List-(\d{4}-\d{2}-\d{2}-\d{2}-\d{2}-\d{2})\.xlsx$',
    re.IGNORECASE,
)


def _find_latest_blob(container_client) -> str | None:
    candidates = []
    for blob in container_client.list_blobs():
        m = _BLOB_NAME_RE.match(blob.name)
        if m:
            dt = datetime.strptime(m.group(1), '%Y-%m-%d-%H-%M-%S')
            candidates.append((dt, blob.name))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


def _download_blob(container_client, blob_name: str, dest_path: str) -> None:
    blob_client = container_client.get_blob_client(blob_name)
    with open(dest_path, 'wb') as f:
        blob_client.download_blob().readinto(f)


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------
def _detect_format(ws) -> str:
    """Return 'sfdc' or 'delaware' based on the value of cell A1."""
    val = ws.cell(1, 1).value
    if val and str(val).strip() == _SFDC_SENTINEL:
        return 'sfdc'
    return 'delaware'


# ---------------------------------------------------------------------------
# Delaware "As of" date parsing
# ---------------------------------------------------------------------------
_AS_OF_RE = re.compile(
    r'As\s+of\s+(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})',
    re.IGNORECASE,
)


def _parse_as_of(ws) -> datetime | None:
    cell_val = ws.cell(_DEL_AS_OF_ROW, _DEL_AS_OF_COL).value or ''
    m = _AS_OF_RE.search(str(cell_val))
    if m:
        return datetime.strptime(m.group(1), '%Y-%m-%d %H:%M:%S')
    return None


# ---------------------------------------------------------------------------
# Delaware address parsing
# Splits known country names off the end of the raw address string.
# ---------------------------------------------------------------------------
_COUNTRIES = [
    'United States', 'United Kingdom', 'Canada', 'Singapore',
    'Guernsey', 'Jersey', 'Cayman Islands', 'British Virgin Islands',
    'Germany', 'France', 'Netherlands', 'Luxembourg', 'Switzerland',
    'Australia', 'Japan', 'China', 'Hong Kong', 'Ireland',
]


def _parse_address(raw: str) -> dict:
    if not raw:
        return {}
    addr = re.sub(r'[\r\n]+', ' ', raw).strip()
    addr = re.sub(r'\s{2,}', ' ', addr)
    if addr in _COUNTRIES:
        return {'COUNTRY': addr}
    country = None
    for ctry in _COUNTRIES:
        if addr.endswith(ctry):
            country = ctry
            addr = addr[: -len(ctry)].rstrip(', ')
            break
    return {'ADDRESS1': addr or None, 'COUNTRY': country}


# ---------------------------------------------------------------------------
# Row parsers
# ---------------------------------------------------------------------------
def _s(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _parse_delaware_rows(ws, as_of: datetime | None) -> list:
    rows = []
    for raw_row in ws.iter_rows(min_row=_DEL_DATA_START, values_only=True):
        org_name   = raw_row[_DEL_ORG_COL   - 1]
        first_name = raw_row[_DEL_FNAME_COL  - 1]
        last_name  = raw_row[_DEL_LNAME_COL  - 1]
        addr_raw   = raw_row[_DEL_ADDR_COL   - 1]

        if not org_name and not first_name and not last_name:
            continue

        addr = _parse_address(str(addr_raw) if addr_raw else '')
        rows.append({
            'Upload_Date': as_of,
            'ORG_NAME':    _s(org_name),
            'FIRST_NAME':  _s(first_name),
            'LAST_NAME':   _s(last_name),
            'ADDRESS1':    addr.get('ADDRESS1'),
            'CITY':        None,
            'STATE':       None,
            'POSTAL_CODE': None,
            'COUNTRY':     addr.get('COUNTRY'),
            'Contact_ID':  None,
            'Entity_ID':   None,
        })
    return rows


def _parse_sfdc_rows(ws, upload_date: datetime) -> list:
    rows = []
    for raw_row in ws.iter_rows(min_row=_SFDC_DATA_START, values_only=True):
        contact_id = _s(raw_row[_SFDC_CONTACT_ID - 1])
        entity_id  = _s(raw_row[_SFDC_ENTITY_ID  - 1])
        org_name   = _s(raw_row[_SFDC_ORG_NAME   - 1])
        first_name = _s(raw_row[_SFDC_FIRST_NAME  - 1])
        last_name  = _s(raw_row[_SFDC_LAST_NAME   - 1])
        street     = _s(raw_row[_SFDC_STREET      - 1])
        city       = _s(raw_row[_SFDC_CITY        - 1])
        state      = _s(raw_row[_SFDC_STATE       - 1])
        postal     = _s(raw_row[_SFDC_POSTAL      - 1])
        country    = _s(raw_row[_SFDC_COUNTRY     - 1])

        if not contact_id and not entity_id and not org_name and not first_name and not last_name:
            continue

        rows.append({
            'Upload_Date': upload_date,
            'ORG_NAME':    org_name,
            'FIRST_NAME':  first_name,
            'LAST_NAME':   last_name,
            'ADDRESS1':    street,
            'CITY':        city,
            'STATE':       state,
            'POSTAL_CODE': postal,
            'COUNTRY':     country,
            'Contact_ID':  contact_id,
            'Entity_ID':   entity_id,
        })
    return rows


# ---------------------------------------------------------------------------
# Database insert
# ---------------------------------------------------------------------------
def _insert_rows(conn, schema: str, rows: list) -> None:
    sql = f"""
        INSERT INTO [{schema}].[ScreeningInput]
            (Upload_Date,
             ORG_NAME, FIRST_NAME, LAST_NAME,
             ADDRESS1, CITY, STATE, POSTAL_CODE, COUNTRY,
             Contact_ID, Entity_ID)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """
    params = [
        (r['Upload_Date'],
         r['ORG_NAME'],   r['FIRST_NAME'], r['LAST_NAME'],
         r['ADDRESS1'],   r['CITY'],       r['STATE'],
         r['POSTAL_CODE'],r['COUNTRY'],
         r['Contact_ID'], r['Entity_ID'])
        for r in rows
    ]
    cur = conn.cursor()
    cur.fast_executemany = True
    cur.executemany(sql, params)
    conn.commit()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(
        description='Import screening list xlsx into ScreeningInput'
    )
    ap.add_argument('--input',     default=None, metavar='PATH',
                    help='Local xlsx file. Omit to auto-discover Delaware OFAC List from blob.')
    ap.add_argument('--container', default=os.environ.get('STORAGE_CONTAINER', 'sdn'),
                    metavar='CONTAINER')
    ap.add_argument('--server',   default=os.environ.get('SQL_SERVER'),   metavar='FQDN')
    ap.add_argument('--database', default='SDN', metavar='DB')
    ap.add_argument('--schema',   default='dbo', metavar='SCHEMA')
    ap.add_argument('--user',     default=os.environ.get('SQL_USER'),     metavar='USER')
    ap.add_argument('--password', default=os.environ.get('SQL_PASSWORD'), metavar='PWD')
    ap.add_argument('--drop', action='store_true',
                    help='TRUNCATE ScreeningInput before importing')
    args = ap.parse_args()

    server   = args.server   or os.environ.get('SQL_SERVER')
    user     = args.user     or os.environ.get('SQL_USER')
    password = args.password or os.environ.get('SQL_PASSWORD')
    if not server:
        sys.exit('ERROR: --server not set (or SQL_SERVER env var missing)')

    # ---- Resolve input file --------------------------------------------------
    tmp_path = None
    if args.input:
        xlsx_path = args.input
    else:
        conn_str = os.environ.get('STORAGE_CONNECTION_STRING')
        if not conn_str:
            sys.exit('ERROR: --input not provided and STORAGE_CONNECTION_STRING is not set.')
        try:
            from azure.storage.blob import BlobServiceClient
        except ImportError:
            sys.exit('ERROR: azure-storage-blob not installed.')

        print(f'Scanning container "{args.container}" for Delaware OFAC List-*.xlsx...',
              flush=True)
        container_client = (BlobServiceClient
                            .from_connection_string(conn_str)
                            .get_container_client(args.container))
        blob_name = _find_latest_blob(container_client)
        if not blob_name:
            print(f'  No matching file found in "{args.container}". '
                  'ScreeningInput unchanged — skipping import.')
            return

        print(f'  Found: {blob_name}', flush=True)
        tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(tmp_fd)
        print('  Downloading...', flush=True)
        _download_blob(container_client, blob_name, tmp_path)
        xlsx_path = tmp_path

    # ---- Load workbook -------------------------------------------------------
    rows = []
    try:
        print(f'Loading {xlsx_path}...', flush=True)
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        fmt = _detect_format(ws)
        print(f'  Format detected: {fmt}', flush=True)

        if fmt == 'sfdc':
            upload_date = datetime.utcnow()
            print(f'  Upload date (UTC): {upload_date.strftime("%Y-%m-%d %H:%M:%S")}',
                  flush=True)
            rows = _parse_sfdc_rows(ws, upload_date)

        else:  # delaware
            as_of = _parse_as_of(ws)
            if as_of:
                print(f'  As of: {as_of}', flush=True)
            else:
                print('  WARNING: Could not parse "As of" date — Upload_Date will be NULL')
            rows = _parse_delaware_rows(ws, as_of)

        wb.close()
        print(f'  {len(rows):,} rows parsed', flush=True)

    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)

    if not rows:
        print('  No data rows found — nothing to import.')
        return

    # ---- Connect and insert --------------------------------------------------
    cs = (
        f'DRIVER={{ODBC Driver 17 for SQL Server}};'
        f'SERVER={server};DATABASE={args.database};'
        f'UID={user};PWD={password};'
        'Encrypt=yes;TrustServerCertificate=no;Connection Timeout=30;'
    )
    if not user:
        cs = (
            f'DRIVER={{ODBC Driver 17 for SQL Server}};'
            f'SERVER={server};DATABASE={args.database};'
            'Trusted_Connection=yes;'
        )

    print(f'Connecting to [{server}].[{args.database}]...', flush=True)
    with pyodbc.connect(cs) as conn:
        exists = conn.cursor().execute(
            "SELECT 1 FROM INFORMATION_SCHEMA.TABLES "
            "WHERE TABLE_SCHEMA=? AND TABLE_NAME='ScreeningInput'",
            [args.schema]
        ).fetchone()
        if not exists:
            sys.exit(
                'ERROR: ScreeningInput does not exist yet.\n'
                'Run sdn_match_v2.py once (with --setup) to create the table first.'
            )

        if args.drop:
            conn.cursor().execute(f'TRUNCATE TABLE [{args.schema}].[ScreeningInput]')
            conn.commit()
            print('  ScreeningInput truncated.', flush=True)

        _insert_rows(conn, args.schema, rows)

    print(f'  {len(rows):,} rows inserted into [{args.schema}].[ScreeningInput]', flush=True)


if __name__ == '__main__':
    main()
